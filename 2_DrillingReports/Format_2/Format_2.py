import openpyxl
import os
import re
import cell_manager
import datetime

result_link = r""
# Название первого блока с данными
current_data_block = r"Эксплуатационное бурение"
result_workbook = openpyxl.Workbook()


def main():
    # Ссылка на папку с файлами для обработки
    link = r""
    print("Processing file \n {} \n\n".format(link))
    global result_link
    result_link = make_result_link(link)
    input_workbook = openpyxl.load_workbook(link, data_only=True)
    clear_input_workbook(input_workbook)
    for worksheet in input_workbook.worksheets:
        print(worksheet)
        # Строка, начиная с которой будут обрабатываться данные
        first_row = 6
        last_row = get_data_length(worksheet, column="AR", stop='Вид бурения') - 2
        try:
            process_sheet(worksheet, first_row, last_row)
        except:
            pass
        result_workbook.save(result_link)
        print(result_link)


def get_data_length(ws, column="A", start_row=1, stop=None):
    """Возвращает номер последней непустой строки в указанном столбце

    :param ws: Страница excel, работа с которой ведется
    :param column: Столбец, количество строк в котором определяется
    :param start_row: Строка, с которой начнется счет
    :param stop:
    :return:
    """
    rows = ws.max_row
    start_cell = "{}{}".format(column, start_row)
    end_cell = "{}{}".format(column, rows)
    interval = ws[start_cell: end_cell]
    # Нас интересует промежуток данных, до слова "Показатели" в столбце A
    for row in reversed(interval):
        for item in row:
            if item.value == stop:
                rows -= 1
                return rows
            else:
                rows -= 1


def process_well_cell(cell):
    """ Очищает имя скважины из ячейки, для того, чтобы можно было создать страницу excel с нужным названием


    :param cell: Значение ячейки, содержащей название скважины
    :return: Очищенное название скважины
    """
    value = cell.value
    value = str(value)
    index = value.find("\n")
    cleared_cell_value = value[:index]
    trash = [r"\\", "/", ":", "?", "*", "[", "]"]
    for icon in trash:
        cleared_cell_value = cleared_cell_value.replace(icon, "")
    return cleared_cell_value


def process_sheet(input_worksheet, first_row, last_row):
    """ Функция получает на вход лист для обработки, и интервал строк, который необходимо обработать, на выходе отдает
    словарь, содержащий данные из этого интервала


    :param input_worksheet: Лист Excel, обработка которого будет производиться
    :param first_row: Строка, начиная с которой будет идти обработка данных
    :param last_row: Строка, начиная до которой будет идти обработка данных
    :return: current_row_data - Словарь, содержащий в качестве ключей названия ячеек,
    в качестве значений - содержащиеся в них данные, для текущего листа Excel
    """
    # Формат ячейки, в которой указывается название скважины
    well_cell_pattern = r"^F[\d]+"
    # Названия существующих на данный момент блоков с данными по скважинам
    list_of_sections = ["Эксплуатационное бурение",
                        "Опережающее бурение",
                        "Геолого-разведочное бурение (ГРР)",
                        "Реконструкция скважин (ЗБС)",
                        "Освоение, испытание скважин"]
    row_number = first_row
    global current_data_block
    while row_number < last_row:
        status_cell = "A{}".format(row_number)
        status_cell_value = str(input_worksheet[status_cell].value).strip()
        # Если происходит смена блока, запоминаем название нового, и пропускаем все скрытые строки
        if status_cell_value in list_of_sections:
            current_data_block = status_cell_value
            row_number += 1
            while input_worksheet.row_dimensions[row_number].hidden:
                row_number += 1

        current_row_data = {}
        row_in_process = input_worksheet[row_number]

        # Проверка, находится ли скважина в бурении, и не скрыта ли эта строка (значит уже не в бурении)
        if not input_worksheet.row_dimensions[row_number].hidden:
            for cell in row_in_process:
                try:
                    cell_column = cell.column_letter
                    cell_name = cell_column + str(row_number)
                    if re.match(well_cell_pattern, cell_name):
                        cleared_cell_value = process_well_cell(cell)
                        current_row_data[cell_name] = cleared_cell_value
                    else:
                        # Если не ячейка с названием скважины, то при помощи cell_manager определяется какая это ячейка,
                        # и как с ней необходимо работаь
                        current_row_data[cell_name] = cell_manager.manage_cells(worksheet=input_worksheet,
                                                                                cell_name=cell_name,
                                                                                row_number=row_number)
                # Не чекал когда возникает, возможно, стоит убрать исключение, и посмотреть, что будет
                except AttributeError:
                    pass
            make_worksheets(current_row_data)
            sheet_name = input_worksheet.title
            try:
                write_data(current_row_data, sheet_name)
            except UnboundLocalError:
                pass
        # На каждую скважину приходится 4 объединенных ячейки, поэтому при переходе со скв на скв шаг - 4 строки
        step = 4
        row_number += step


def make_worksheets(current_row_data):
    """ На основании данных из строк создает листы с именами скважин в результирующем файле


    :param current_row_data:
    :return:
    """
    # Ячейка, в которой содержится название нужной скважины
    well_cell_pattern = r"^F[\d]+"
    for cell_name, value in current_row_data.items():
        value = str(value)
        if re.match(well_cell_pattern, cell_name):
            if value not in result_workbook.sheetnames:
                worksheet = result_workbook.create_sheet(value)
                make_hat(worksheet)
            else:
                pass


def validate_value(value):
    """Функция получает на вход значение, проверяет, является ли оно числом, датой, None, или строкой, и форматирует
    нужным образом

    :param value: Значение
    :return: Отформатированное значение
    """

    if value is None:
        value = "-"
    else:
        try:
            int(value)
        except TypeError:
            value = str(value)
            datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S').date()
            value = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S').date()
            value = value.strftime("%d.%m.%Y")
        except ValueError:
            value = str(value)
    return value


def write_data(current_row_data, sheet_name):
    """Функция заносит данные на необходимую страничку


    :param current_row_data: Данные для переноса
    :param sheet_name: Название листа, с которого брались данные, нужно для заполнения даты
    :return:
    """
    global current_data_block
    # Словарь, используемый далее в цикле для определения данные из каких столбцов в какие заносить, ключи - из
    # каких столбцов, значения - в какие столбцы
    dict_to_put = {"C": "C",
                   "H": "D",
                   "F": "E",
                   "J": "G",
                   "K": "I",
                   "G": "K",
                   "N": "L",
                   "I": "N",
                   "AD": "P",
                   "AR": "Q",
                   "BF": "R"
                   }

    # Блок для нахождения листа Excel, отвечающего за определенную скважину
    well_cell_pattern = r"^F[\d]+"
    for cell_name, value in current_row_data.items():
        value = str(value)
        if re.match(well_cell_pattern, cell_name):
            worksheet = result_workbook[value]
    row_to_write = worksheet.max_row + 1

    # Блок заноса данных в листы Excel, поиск подходящего столбца сделан через соответствия названия ячейки, из которой
    # были взяты данные в исходном файле, определенному регулярному выражению
    for key, value in current_row_data.items():
        # Прописывается дата в столбце А, статус скважины - в столбце B
        sheet_name = str(sheet_name)
        sheet_name = sheet_name.replace('-', '.')
        worksheet["A{}".format(row_to_write)].value = sheet_name
        worksheet["B{}".format(row_to_write)].value = current_data_block

        # Заносятся остальные данные, кроме даты
        value = str(value)
        for where_from, where_to in dict_to_put.items():
            well_cell_pattern = r"^{}[\d]+".format(where_from)
            if re.match(well_cell_pattern, key):
                # Производится приведение value к нужному типу данных (кажись не арбайтен)
                value = validate_value(value)
                worksheet["{}{}".format(where_to, row_to_write)].value = value
                continue


def make_result_link(input_link):
    """Создает ссылку на результирующий файл на основании входного


    :param input_link: Ссылка на файл для обработки
    :return: Ссылка на результирующий файл
    """
    folder_name = os.path.dirname(input_link)
    filename = "Result.xlsx"
    result_link = os.path.join(folder_name, filename)
    return result_link


def make_hat(worksheet):
    """Прописывает шапку для вывода информации


    :param worksheet: Лист Excel, на котором будет создаваться шапка
    :return:
    """
    global result_workbook
    # Необходимо указать ячейки и что в них написать
    worksheet["A2"] = "Дата"
    worksheet["B2"] = "Назначение скважины"
    worksheet["C2"] = "Бур.подрядчик"
    worksheet["D2"] = "Тип БУ №"
    worksheet["E2"] = "№ куста/ДГУ заказ. (+,-)"
    worksheet["F2"] = "№ скв."
    worksheet["G2"] = "Статус "
    worksheet["H2"] = "План м3/сут, План.  т/сут ГТМ"
    worksheet["I2"] = "Начало факт"
    worksheet["J2"] = "Сдача устья план ГТМ"
    worksheet["K2"] = "Сдача устья прогноз"
    worksheet["L2"] = "отклонение Сдача устья прогноз "
    worksheet["M2"] = "Наличие БРД, КРС на кусту "
    worksheet["N2"] = "Конструкция план"
    worksheet["O2"] = "Конструкция факт"
    worksheet["P2"] = "забой на начало суток"
    worksheet["Q2"] = "забой на конец суток"
    worksheet["R2"] = "Проходка за сутки факт"
    worksheet["S2"] = "Проходка за сутки план"
    worksheet["T2"] = "Отклонение"
    worksheet["U2"] = "План на текущие сутки 00:00-24:00 "
    worksheet["V2"] = "+/- относительно графика 'Глубина/день'"
    worksheet["W2"] = "НПВ более 6ч."
    worksheet["X2"] = "Баланс времени, час"
    result_workbook.save(result_link)


def clear_input_workbook(input_workbook):
    """ Функция удаляет страницы с определенными названиями из результирующего файла

    :param input_workbook:
    :return:
    """
    forbidden_sheets = ["Sheet",
                        "конструкции",
                        "30.08.20"]
    for sheet in forbidden_sheets:
        if sheet in input_workbook.sheetnames:  # remove default sheet
            input_workbook.remove(input_workbook[sheet])


if __name__ == '__main__':
    main()
