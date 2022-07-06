import openpyxl
import os
import re
import datetime
import multiprocessing


# Необходимые глобальные переменные
result_link = r""
result_workbook = openpyxl.Workbook()


def multiprocess():
    """Создает пул процессов, производит обработку
    """
    # Необходимо указать ссылку на папку с файлами для обработки
    link = r""
    workbooks_list = make_list_of_files(link)
    with multiprocessing.Pool() as pool:
        pool.map(main, workbooks_list)


def main(link):
    print("Processing folder \n {} \n\n".format(link))
    global result_link
    result_link = make_result_link(link)
    input_workbook = openpyxl.load_workbook(link, data_only=True)
    for worksheet in input_workbook.worksheets:
        print(worksheet)
        # Строка, начиная с которой будут обрабатываться данные
        first_row = 8
        last_row = get_data_length(worksheet, stop="Показатели")
        process_sheet(worksheet, first_row, last_row)
        result_workbook.save(result_link)


def get_data_length(ws, column="A", start_row=1, stop=None):
    """Возвращает номер последней непустой строки в указанном столбце

    :param ws: Страница excel, работа с которой ведется
    :param column: Столбец, количество строк в котором определяется
    :param start_row: Строка, с которой начнется счет
    :param stop:
    :return:
    """
    rows = ws.max_row
    start_cell = "{}{}".format(column, start_row)
    end_cell = "{}{}".format(column, rows)
    interval = ws[start_cell: end_cell]
    # Нас интересует промежуток данных, до слова "Показатели" в столбце A
    for row in reversed(interval):
        for item in row:
            if item.value == stop:
                rows -= 1
                return rows
            else:
                rows -= 1


def process_sheet(input_worksheet, first_row, last_row):
    """ Функция получает на вход лист для обработки, и интервал строк, который необходимо обработать, на выходе отдает
    словарь, содержащий данные из этого интервала


    :param input_worksheet: Лист Excel, обработка которого будет производиться
    :param first_row: Строка, начиная с которой будет идти обработка данных
    :param last_row: Строка, начиная до которой будет идти обработка данных
    :return: current_row_data - Словарь, содержащий в качестве ключей названия ячеек,
    в качестве значений - содержащиеся в них данные, для текущего листа Excel
    """
    for row_number in range(first_row, last_row):
        current_row_data = {}
        row_in_process = input_worksheet[row_number]
        # Ячейка - в которой указывается статус бурения скважины
        well_status_cell = "J{}".format(row_number)
        # Проверка, находится ли скважина в бурении, и не скрыта ли эта строка (значит уже не в бурении)
        if input_worksheet[well_status_cell].value == "Бурение" and not input_worksheet.row_dimensions[row_number].hidden:
            for cell in row_in_process:
                try:
                    cell_column = cell.column_letter
                    cell_name = cell_column + str(row_number)
                    current_row_data[cell_name] = cell.value
                except AttributeError:
                    pass
            make_worksheets(current_row_data)
            sheet_name = input_worksheet.title
            write_data(current_row_data, sheet_name)


def make_worksheets(current_row_data):
    """ На основании данных из строк создает листы с именами скважин в результирующем файле


    :param current_row_data:
    :return:
    """
    well_cell_pattern = r"^I[\d]+"
    for cell_name, value in current_row_data.items():
        value = str(value)
        if re.match(well_cell_pattern, cell_name):
            if value not in result_workbook.sheetnames:
                worksheet = result_workbook.create_sheet(value)
                make_hat(worksheet)
            else:
                pass


def validate_value(value):
    """Функция получает на вход значение, проверяет, является ли оно числом, датой, None, или строкой, и форматирует
    нужным образом

    :param value: Значение
    :return: Отформатированное значение
    """

    if value is None:
        value = "-"
    else:
        try:
            int(value)
        except TypeError:
            value = str(value)
            datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S').date()
            value = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S').date()
            value = value.strftime("%d.%m.%Y")
        except ValueError:
            value = str(value)
    return value


def write_data(current_row_data, sheet_name):
    """Функция заносит данные на необходимую страничку


    :param current_row_data: Данные для переноса
    :param sheet_name: Название листа, с которого брались данные, нужно для заполнения даты
    :return:
    """
    # Словарь, используемый далее в цикле для определения данные из каких столбцов в какие заносить, ключи - из
    # каких столбцов, значения - в какие столбцы
    dict_to_put = {"A": "B",
                   "D": "C",
                   "E": "D",
                   "H": "E",
                   "I": "F",
                   "J": "G",
                   "K": "H",
                   "L": "I",
                   "M": "J",
                   "N": "K",
                   "O": "L",
                   "P": "M",
                   "Q": "N",
                   "R": "O",
                   "S": "P",
                   "T": "Q",
                   "U": "R",
                   "V": "S",
                   "W": "T",
                   "X": "U",
                   "Y": "V",
                   "AM": "W",
                   "AN": "X",
                   }
    # Блок для нахождения листа Excel, отвечающего за определенную скважину
    well_cell_pattern = r"^I[\d]+"
    for cell_name, value in current_row_data.items():
        if re.match(well_cell_pattern, cell_name):
            value = str(value)
            worksheet = result_workbook[value]
            print(worksheet)
    row_to_write = worksheet.max_row + 1

    # Блок заноса данных в листы Excel, поиск подходящего столбца сделан через соответствия названия ячейки, из которой
    # были взяты данные в исходном файле, определенному регулярному выражению
    for key, value in current_row_data.items():
        # Прописывается дата в столбце А, нужно допилить форматирование даты
        sheet_name = str(sheet_name)
        sheet_name = sheet_name.replace('-', '.')
        worksheet["A{}".format(row_to_write)].value = sheet_name
        # Заносятся остальные данные, кроме даты
        for where_from, where_to in dict_to_put.items():
            well_cell_pattern = r"^{}[\d]+".format(where_from)
            if re.match(well_cell_pattern, key):
                value = validate_value(value)
                worksheet["{}{}".format(where_to, row_to_write)].value = value
                continue


def make_result_link(input_link):
    """Создает ссылку на результирующий файл на основании входного


    :param input_link: Ссылка на файл для обработки
    :return: Ссылка на результирующий файл
    """
    result_folder = "./2_output/"
    filename = os.path.basename(input_link)
    # filename = "Result.xlsx"
    result_link = os.path.join(result_folder, filename)
    return result_link


def make_hat(worksheet):
    """Прописывает шапку для вывода информации


    :param worksheet: Лист Excel, на котором будет создаваться шапка
    :return:
    """
    global result_workbook
    # Необходимо указать ячейки и что в них написать
    worksheet["A2"] = "Дата"
    worksheet["B2"] = "Назначение скважины"
    worksheet["C2"] = "Бур.подрядчик"
    worksheet["D2"] = "Тип БУ №"
    worksheet["E2"] = "№ куста/ДГУ заказ. (+,-)"
    worksheet["F2"] = "№ скв."
    worksheet["G2"] = "Статус "
    worksheet["H2"] = "План м3/сут, План.  т/сут ГТМ"
    worksheet["I2"] = "Начало факт"
    worksheet["J2"] = "Сдача устья план ГТМ"
    worksheet["K2"] = "Сдача устья прогноз"
    worksheet["L2"] = "отклонение Сдача устья прогноз "
    worksheet["M2"] = "Наличие БРД, КРС на кусту "
    worksheet["N2"] = "Конструкция план"
    worksheet["O2"] = "Конструкция факт"
    worksheet["P2"] = "забой на начало суток"
    worksheet["Q2"] = "забой на конец суток"
    worksheet["R2"] = "Проходка за сутки факт"
    worksheet["S2"] = "Проходка за сутки план"
    worksheet["T2"] = "Отклонение"
    worksheet["U2"] = "План на текущие сутки 00:00-24:00 "
    worksheet["V2"] = "+/- относительно графика 'Глубина/день'"
    worksheet["W2"] = "НПВ более 6ч."
    worksheet["X2"] = "Баланс времени, час"
    result_workbook.save(result_link)


def make_list_of_files(link):
    """Создает лист со ссылками на необходимые файлы, необходим для многопроцессорной обработки


    :param link: Ссылка на папку, внутри которой необходимо будет искать файлы
    :return:
    """
    links_list = []
    for root, directories, files in os.walk(link):
        for file in files:
        tag = ""
            if tag in file:
                file_path = os.path.join(root, file)
                links_list.append(file_path)
    return links_list

if __name__ == '__main__':
    multiprocess()
