import openpyxl
from openpyxl.chart import (LineChart, Reference)


def main():
    # Ссылка на Excel файл для обработки
    link = r""
    workbook = openpyxl.load_workbook(link)
    clear_input_workbook(workbook)
    for worksheet in workbook.worksheets:
        print("Processing worksheet {}".format(worksheet))
        new_table_start = worksheet.max_row + 5
        last_row = make_calculations(worksheet, new_table_start)
        make_graph(worksheet, new_table_start, last_row)
    workbook.save(link)


def make_calculations(worksheet, new_table_start):
    """Функция производит рассчет накопленной проходки


    :param worksheet:
    :param new_table_start:
    :return:
    """
    max_row = worksheet.max_row
    header_line = 2
    worksheet["B{}".format(new_table_start + 3)].value = "Забой на конец суток факт"
    worksheet["C{}".format(new_table_start + 3)].value = "Забой на конец суток план"
    worksheet["B{}".format(new_table_start + 4)].value = worksheet["R{}".format(header_line + 1)].value
    worksheet["C{}".format(new_table_start + 4)].value = worksheet["S{}".format(header_line + 1)].value
    for i in range(header_line + 3, max_row + header_line):
        zaboy_last_day_fact = worksheet["B{}".format(new_table_start + i - 1)].value
        zaboy_last_day_plan = worksheet["C{}".format(new_table_start + i - 1)].value
        prohodka_fact = worksheet["R{}".format(i - 1)].value
        prohodka_plan = worksheet["S{}".format(i - 1)].value

        if zaboy_last_day_fact is None:
            zaboy_last_day_fact = 0
        if zaboy_last_day_plan is None:
            zaboy_last_day_plan = 0
        if prohodka_fact is None:
            prohodka_fact = 0
        if prohodka_plan is None:
            prohodka_plan = 0

        zaboy_last_day_fact = int(zaboy_last_day_fact)
        zaboy_last_day_plan = int(zaboy_last_day_plan)
        prohodka_fact = int(prohodka_fact)
        prohodka_plan = int(prohodka_plan)
        worksheet["A{}".format(new_table_start + i - 1)].value = worksheet["A{}".format(i - 2)].value
        worksheet["B{}".format(new_table_start + i)].value = zaboy_last_day_fact + prohodka_fact
        worksheet["C{}".format(new_table_start + i)].value = zaboy_last_day_plan + prohodka_plan

    worksheet["A{}".format(new_table_start + i)].value = worksheet["A{}".format(i - 1)].value
    last_row = new_table_start + i
    return last_row


def make_graph(worksheet, new_table_start, last_row):
    """ Функция производит построение графиков


    :param worksheet: Лист Excel, на котором будет построен график
    :param new_table_start: Первая строка с данными
    :param last_row: Последняя строка с данными
    :return:
    """
    chart_1 = LineChart()
    chart_1.title = "График бурения скважины"
    chart_1.style = 13
    chart_1.x_axis.title = "Дата"
    chart_1.legend.position = "b"
    chart_1.y_axis.title = "Глубина скважины"
    chart_1.y_axis.scaling.orientation = "maxMin"
    chart_1.height = 15
    chart_1.width = 20

    data = Reference(worksheet, min_col=2, min_row=new_table_start + 3, max_col=3, max_row=last_row)
    chart_1.add_data(data, titles_from_data=True)
    x_values = Reference(worksheet, min_col=1, min_row=new_table_start + 4, max_row=last_row)

    chart_1.set_categories(x_values)

    style_fact_line = chart_1.series[0]
    style_fact_line.marker.symbol = "circle"
    style_fact_line.marker.size = 14
    style_fact_line.marker.graphicalProperties.solidFill = "6699FF"
    style_fact_line.marker.graphicalProperties.line.solidFill = "6699FF"
    style_fact_line.graphicalProperties.line.solidFill = "6699FF"

    style_plan_line = chart_1.series[1]
    style_plan_line.marker.symbol = "circle"
    style_plan_line.marker.size = 14
    style_plan_line.marker.graphicalProperties.solidFill = "CC0000"
    style_plan_line.marker.graphicalProperties.line.solidFill = "CC0000"
    style_plan_line.graphicalProperties.line.solidFill = "CC0000"

    worksheet.add_chart(chart_1, "E{}".format(new_table_start + 3))


def clear_input_workbook(input_workbook):
    """ Функция удаляет страницы с определенными названиями из результирующего файла

    :param input_workbook:
    :return:
    """
    forbidden_sheets = ["Sheet",
                        "конструкции",
                        "30.08.20"]
    for sheet in forbidden_sheets:
        if sheet in input_workbook.sheetnames:  # remove default sheet
            input_workbook.remove(input_workbook[sheet])


if __name__ == '__main__':
    main()
