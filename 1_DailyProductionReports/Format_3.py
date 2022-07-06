import os
import sys
import openpyxl
import re
import win32com.client as win32


def main(folder_to_process, result_file):
    result_workbook = openpyxl.load_workbook(result_file)
    result_worksheet = result_workbook.worksheets[0]

    # Получить список необходимых дат
    dates_list = process_dates(result_worksheet)
    row_number = get_data_length(result_worksheet, column="C", start_row=1) + 1

    preprocess_files(folder_to_process)
    for date in dates_list:
        day, month = get_day_month(date)
        path_to_file = form_link(folder_to_process, day, month)
        try:
            svodka_book = openpyxl.load_workbook(path_to_file, data_only=True)
            svodka_vstavka_sheet = svodka_book["SheetName1"]
            svodka_fond_sheet = svodka_book["SheetName2"]
            vstavka_values = take_vstavka_values(svodka_vstavka_sheet)
            put_vstavka_values(result_worksheet, row_number, vstavka_values)
            fond_values = take_fond_values(svodka_fond_sheet)
            put_fond_values(result_worksheet, row_number, fond_values)

        except:
            print("Problem with file: ", path_to_file)
            print(sys.exc_info())
            pass
        row_number += 1
        result_workbook.save(result_file)


# Словарь, используемый в последующих фукнциях, ключи словаря - ячейки из сводок, которые понадобятся в работе,
# значения словаря - в какие столбцы в реультате надо будет вставить значения
vstavka_columns_dict = {"D1": "A",
                        "D17": "C",
                        "D14": "D",
                        "D23": "E",
                        "D15": "F",
                        "D5": "G",
                        "D2": "H",
                        "D8": "I",
                        "D3": "J",
                        "D50": "K",
                        "D47": "L",
                        "D53": "M",
                        "D48": "N",
                        "D59": "O",
                        "D56": "P",
                        "D62": "Q",
                        "D29": "R",
                        "D32": "S",
                        "D44": "T"
                        }

fond_columns_dict = {"B5": "U",
                     "B6": "V",
                     "B7": "W",
                     "B8": "X",
                     "B9": "Y",
                     "B10": "Z",
                     "B13": "AA"
                     }


def get_data_length(ws, column="A", start_row=1):
    """Возвращает номер последней непустой строки в указанном столбце

    :param ws: Страница excel, работа с которой ведется
    :param column: Столбец, количество строк в котором определяется
    :param start_row: Строка, с которой начнется счет
    :return:
    """
    rows = ws.max_row
    start_cell = "{}{}".format(column, start_row)
    end_cell = "{}{}".format(column, rows)
    interval = ws[start_cell: end_cell]
    for row in reversed(interval):
        for item in row:
            if item.value is None:
                rows -= 1
            else:
                return rows


def take_vstavka_values(worksheet):
    """Функция считывает значения из ячеек в list_of_cells, и сохраняет их в cell_values


    :param worksheet: Страница Excel, с которой берутся данные
    :return: Словарь, содержащий названия ячеек и значения в них
    """

    cells_values = {}
    for cell, column in vstavka_columns_dict.items():
        cells_values[column] = worksheet[cell].value
    return cells_values


def take_fond_values(worksheet):
    """Функция считывает значения из ячеек в list_of_cells, и сохраняет их в cell_values


    :param worksheet: Страница Excel, с которой берутся данные
    :return: Словарь, содержащий названия ячеек и значения в них
    """

    cells_values = {}
    for cell, column in fond_columns_dict.items():
        cells_values[column] = worksheet[cell].value
    return cells_values


def put_vstavka_values(result_worksheet, row_number, values):
    """Функция получает на вход словарь со значениями из сводки,
    и вставляет их в необходимые столбцы result_worksheet

    :param result_worksheet: Страница excel, в которую необходимо вставить данные
    :param row_number: Номер строки, в которую необходимо вставлять данные
    :param values: Словарь со значениями, которые необходимо вставлять
    :return:
    """
    # Лист столбцов, в которые будет необходимо вставлять данные

    for column, value in values.items():
        output_cell = "{}{}".format(column, row_number)
        result_worksheet[output_cell].value = value


def put_fond_values(result_worksheet, row_number, values):
    """Функция получает на вход словарь со значениями из сводки,
    и вставляет их в необходимые столбцы result_worksheet

    :param result_worksheet: Страница excel, в которую необходимо вставить данные
    :param row_number: Номер строки, в которую необходимо вставлять данные
    :param values: Словарь со значениями, которые необходимо вставлять
    :return:
    """
    # Лист столбцов, в которые будет необходимо вставлять данные

    for column, value in values.items():
        output_cell = "{}{}".format(column, row_number)
        result_worksheet[output_cell].value = value


def process_dates(output_ws):
    """Функция получает на вход страницу excel, в которую нужно поместить значения из сводок, и формирует лист,
    содержащий все даты, данные по которым необходимо собрать


    :param output_ws: Страница excel, в которую будут сводиться данные.
    :return: Лист с необходимыми датами
    """
    dates = []
    last_date_cell = get_data_length(output_ws, column="B", start_row=4) + 1
    for cell in range(4, last_date_cell):
        if output_ws["C{}".format(cell)].value is None:
            date_cell = "B{}".format(cell)
            date = str(output_ws[date_cell].value)
            space_index = date.find(" ")
            date = date[:space_index]
            dates.append(date)
    return dates


def get_day_month(date):
    """Функция при помощи регулярных выражений достет из даты формата YYYY-MM-DD день и месяц


    :param date: Дата в формате YYYY-MM-DD
    :return: Числа дня (31) и месяца(01)
    """
    date_pattern = r"\d{4}[-.](\d{2})[-.](\d{2})"
    date_parts = re.search(date_pattern, date)
    month = date_parts[1]
    day = date_parts[2]
    return day, month


def form_link(link, day, month):
    """На основании ссылки до папки с архивами, формирует ссылку до файла, который необходимо открыть


    :param link: Ссылка на папку с архивами по месяцам
    :param day: День, за который нужна сводка
    :param month: Месяц дня, за который нужна сводка
    :return: Ссылку на файл для дальнейшей обработки
    """
    archive_folders_list = os.listdir(link)
    for archive in archive_folders_list:
        if month in archive:
            archive_link = os.path.join(link, archive)
    day_folders_list = os.listdir(archive_link)
    for day_folder in day_folders_list:
        if day in day_folder:
            folder_link = os.path.join(archive_link, day_folder)

    files = os.listdir(folder_link)
    for file in files:
        try:
            file_tag = r""
            if file_tag in file and file.endswith("xlsx"):
                path_to_file = os.path.join(folder_link, file)
                print(path_to_file)
                return path_to_file
        except UnboundLocalError:
            print("Нет суточной сводки")
            pass


def preprocess_files(link):
    """ Переводит файлы в расшишении .xls в .xlsx


    :param link: Ссылка на папку с файлами для обработки
    :return:
    """
    for root, folders, files in os.walk(link):
        for file in files:
            file_tag = r""
            if file.endswith("xls") and file_tag in file:
                print("I am in folder:\n {}".format(root))
                files_in_folder = os.listdir(root)
                with_new_extension = file.replace(".xls", ".xlsx")
                if with_new_extension not in files_in_folder:
                    print("There is no xlsx file, making one")
                    file_path = os.path.join(root, file)
                    # Переименовываем xls в xlsx
                    excel = win32.gencache.EnsureDispatch('Excel.Application')
                    wb = excel.Workbooks.Open(file_path)
                    wb.SaveAs(file_path + "x", FileFormat=51)
                    wb.Close()
                    excel.Application.Quit()
                    os.remove(file_path)


if __name__ == '__main__':
    # Ссылка на сводки по скважине
    folder_to_process = r""
    # Ссылка, книга, страница, в которую будет вставляться результат
    result_file = r""
    main(folder_to_process, result_file)
