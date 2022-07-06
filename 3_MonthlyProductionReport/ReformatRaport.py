import openpyxl
import time
import re
import os
from tkinter import *
from tkinter import messagebox


def main(source_folder_link, output_folder):

    preprocess_files(source_folder_link)
    result_folder = "{}/MRM_output".format(output_folder)
    os.mkdir(result_folder)
    for root, folders, files in os.walk(source_folder_link):
        for file in files:
            if file.endswith('.xlsx'):
                source_book_link = os.path.join(root, file)
                wellname = get_wellname(source_book_link)
                result_book_link = os.path.join(result_folder, "{}_Output.xlsx".format(wellname))
                process_data(source_book_link, result_book_link)


def run_button_click():
    MRM_folder = Input_text.get()
    output_folder = Output_text.get()
    main(MRM_folder, output_folder)
    messagebox.showinfo('Рассчет окончен', 'Выполнение программы завершено')


# Словарь, используемый для перевода русских обозначений месяцев в цифры
months_library = {"декабрь": "12",
                  "ноябрь": "11",
                  "октябрь": "10",
                  "сентябрь": "09",
                  "август": "08",
                  "июль": "07",
                  "июнь": "06",
                  "май": "05",
                  "апрель": "04",
                  "март": "03",
                  "февраль": "02",
                  "январь": "01"
                  }


def preprocess_files(link):
    """ Переводит файлы в расшишении .xls в .xlsx


    :param link: Ссылка на папку с файлами для обработки
    :return:
    """
    for root, folders, files in os.walk(link):
        for file in files:
            if file.endswith("xls"):
                file_path = os.path.join(root, file)
                # Переименовываем xls в xlsx
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                wb = excel.Workbooks.Open(file_path)
                wb.SaveAs(file_path + "x", FileFormat=51)
                wb.Close()
                excel.Application.Quit()
                os.remove(file_path)


def get_data_length(ws, start_row=1, column="F"):
    """Возвращает номер последней непустой строки в указанном столбце

    :param ws: Страница excel, из которой берутся данные
    :param column: Столбец, количество строк в котором определяется
    :param start_row: Строка, с которой начинается подсчет
    :return:
    """

    rows = ws.max_row
    start_cell = "{}{}".format(column, start_row)
    end_cell = "{}{}".format(column, rows)
    interval = ws[start_cell: end_cell]
    for row in reversed(interval):
        for item in row:
            if item.value is None:
                rows -= 1
            else:
                return rows


def get_chunk_range(source_sheet, start_cell_number, table_top):
    """Функция получает на вход номер стартовой ячейки, и ищет конец данного блока данных (месяца), и возвращает
    значение начала блока

    :param source_sheet: Страница excel, с которой идет работа
    :param start_cell_number: Номер стартовой ячейки обрабатываемого интервала
    :param table_top: Номер ячейки, на которой заканчиваются данные, и начинается шапка
    :return:
    """

    # Строка в столбце F, разделяющая месячные блоки в файле МРМ
    chunk_delimiter = "Состояние"
    while source_sheet["F{}".format(start_cell_number)].value != chunk_delimiter:
        start_cell_number -= 1
        if start_cell_number < table_top:
            break
    start_cell_number += 1
    return start_cell_number


def write_date(result_sheet, source_sheet, start_cell_number):
    """ Функция используется для записи даты в результирующий excel файл.


    :param result_sheet: Страница Excel, в которую вставляются данные
    :param source_sheet: Страница Excel, из которой берутся данные
    :param start_cell_number: Используется для преобразования даты
    :return:
    """
    
    result_sheet["A9"] = "Дата"
    for day in range(7, 38):
        last_data_cell = get_data_length(result_sheet, start_row=9, column="A") + 1
        data_cell = "A{}".format(last_data_cell)
        result_sheet[data_cell].value = make_date(source_sheet, start_cell_number, day)


def make_date(source_sheet, year_month_row, day_date_column):
    """Функция получает на вход две ячейки, в которых содержится информация по дате
       и затем приводит ее к правильному формату


    :param source_sheet: Страница excel, из которой берутся данные
    :param year_month_row: Строка столбца B, из которой берется дата (месяц, год)
    :param day_date_column: Ячейка третьей строки, из которой берется дата (день)
    :return:
    """

    date_day = source_sheet.cell(row=3, column=day_date_column).value
    year_month_cell = "B{}".format(year_month_row)
    date_month_year = source_sheet[year_month_cell].value
    for month in months_library:
        try:
            if month in date_month_year.lower():
                date = format_date(date_day, date_month_year, month)
                return date
        except AttributeError:
            pass


def format_date(date_day, date_month_year, month):
    """Функция получает на вход составные части даты, очищает их, и преобразует


    :param date_day: День
    :param date_month_year: Месяц, год
    :param month: На что необходимо заменить русское название месяца
    :return: Отформатированная дата
    """
    if len(str(date_day)) == 1:
        day_date = "0{}.".format(date_day)
    else:
        day_date = "{}.".format(date_day)
    formatted_date = day_date + date_month_year
    formatted_date = formatted_date.replace(month, months_library[month])
    formatted_date = re.sub(r'[\s]+', '.', formatted_date)
    try:
        time.strptime(formatted_date, "%d.%m.%Y")
        return formatted_date
    # Ошибка возникает если такой даты не существовало
    except ValueError:
        pass


def transpose_values(source_sheet, result_sheet, start_cell_number, end_cell_number):
    """ Функция получает на вход интервал в source_sheet, в котором содержутся данные, из находит среди них нужные, и
        и транспонирует их в result_sheet.


    :param source_sheet: Страница excel, из которой берутся данные
    :param result_sheet: Страница excel, в которую вставляются данные
    :param start_cell_number: Начало интервала обработки (месяц)
    :param end_cell_number: Конец интервала обработки
    :return:
    """
    processed_row_number = start_cell_number

    needed_parameters = {r"Qж, м3/сут": "E",
                         r"Qн, т/сут": "F",
                         r"Обв, %": "H",
                         r"Обв ХАЛ, %": "I",
                         r"Qгаз, м3/сут": "G",
                         r"ГФ, м3/т": "J",
                         r"Рбуф, атм": "Y",
                         r"Рзатр, атм": "Z",
                         r"Рлин, атм": "AA",
                         r"Dшт, мм": "W",
                         r"Рприем, атм ": "AC",
                         r"Рзаб, атм": "P",
                         r"F, Гц": "AE",
                         r"D шт затр, мм": "X"
                         }

    # Написание названий колонок
    for text, letter in needed_parameters.items():
        column_header_cell = "{}9".format(letter)
        result_sheet[column_header_cell] = text
    # Итерация по строкам в интервале start_cell_number:end_cell_number, в столбцах G-AL
    for row in source_sheet.iter_rows(max_col=38, min_row=start_cell_number,
                                      min_col=7, max_row=end_cell_number, values_only=True):
        # Получение названия параметра
        parameter_name = source_sheet["F{}".format(processed_row_number)].value
        # Получение названия параметра
        if parameter_name in needed_parameters:
            column_letter = needed_parameters[parameter_name]
            last_cell = get_data_length(result_sheet, column="{}".format(column_letter))
            for item in row:
                date_cell = "A{}".format(last_cell + 1)
                # Доп условие, чтобы не писались значения, если нет соответствующей им даты
                if result_sheet[date_cell].value is not None:
                    # Транспонировка значений;
                    result_sheet_value_cell = "{}{}".format(column_letter, last_cell + 1)
                    if item is not None:
                        result_sheet[result_sheet_value_cell].value = item
                    else:
                        result_sheet[result_sheet_value_cell].value = "U"
                    last_cell += 1
        processed_row_number += 1


def process_data(source_book_link, result_book_link):
    """ Производит обработку source_book, сохраняет результат в result_book


    :param source_book_link: Ссылка на Excel файл для обработки
    :param result_book_link: Ссылка на Excel файл, в который необходимо сохранить результат
    :return:
    """

    source_book = openpyxl.load_workbook(source_book_link)
    source_sheet = source_book.active

    result_book = openpyxl.Workbook()
    result_sheet = result_book.active

    end_cell_number = get_data_length(source_sheet, start_row=5, column="F")
    start_cell_number = end_cell_number

    table_top = 5

    while start_cell_number >= table_top:

        start_cell_number = get_chunk_range(source_sheet, start_cell_number, table_top)
        write_date(result_sheet, source_sheet, start_cell_number)

        transpose_values(source_sheet, result_sheet, start_cell_number,
                         end_cell_number)
        # -2 необходимо, чтобы перепрыгнуть через ячейку, содержащую ограничитель, который используется для разделения
        # блоков
        end_cell_number = start_cell_number - 2
        start_cell_number = end_cell_number

    result_book.save(result_book_link)


def get_wellname(source_book_link):
    """ Получает ссылку на файл, вовзращает номер скважины

    :param source_book_link: Ссылка на файл
    :return: Номер скважины
    """
    wellname_pattern = r"Скв. ([\d]+)"
    wellname = re.search(wellname_pattern, source_book_link)
    return wellname[1]




# =====================================================================================================================
# ======================================================Интерфейс======================================================
# =====================================================================================================================


# Создание рабочего окна
window = Tk()
window.title("Обработать МРМ")
window.geometry("400x150")

# Условные линии сетки
X_ZERO_LANE = 25
X_FIRST_LANE = 50
X_SECOND_LANE = 50
X_THIRD_LANE = 75


# Создание лейбла способа импорта
import_option_label = Label(window, text="Вставьте ссылку на папку с файлами для обработки:", font=("Arial Bold", 10))
import_option_label.place(x=X_ZERO_LANE, y=0)

# Создание блоков для ввода ссылок
Input_text = Entry(window, width=40)
# Размещение блоков для ввода ссылок
Input_text.place(x=X_SECOND_LANE, y=30)


# Создание лейбла способа импорта
Output_option_label = Label(window, text="В какой папке сохранить результат\n [Папка не должна быть в папке инпута]:"
                            , font=("Arial Bold", 10))
Output_option_label.place(x=X_FIRST_LANE, y=60)

# Создание блоков для ввода ссылок
Output_text = Entry(window, width=40)
# Размещение блоков для ввода ссылок
Output_text.place(x=X_SECOND_LANE, y=100)


gogo_button = Button(window, text="Произвести обработку МРМ", command=run_button_click)
gogo_button.place(x=X_THIRD_LANE, y=120)

window.mainloop()
